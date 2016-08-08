# coding=utf-8
from __future__ import print_function
from future.utils import string_types

from functools import partial
from itertools import product

from collections import Counter
import datetime
import os.path as op

import pandas as pd
import numpy as np
import re

MY_DIR = op.abspath(op.dirname(__file__))
DATA_DIR = op.join(MY_DIR, 'data')
LATEST_XLSX = op.join(DATA_DIR, 'latest.xlsx')


def _trials_to_exp(trials_df,
                   exp_attributes=('exp_id',),
                   series_columns_name='series'):
    # Copied from my pandas utils, docs in there

    exp_attributes = list(exp_attributes)

    trials = [list(trial_metadata) + [trial_df.drop(exp_attributes, axis=1).reset_index(drop=True)]
              for trial_metadata, trial_df in trials_df.groupby(exp_attributes)]

    return pd.DataFrame(trials, columns=exp_attributes + [series_columns_name])


EXPS_NON_STAIRCASE_COLUMNS = [
    'participant',
    'staircase',
    'staircase_repetition',
    'reversal_method',
    'omit',
    'exp_start',
    'age',
    'gender',
    'education',
    'vision',
    'chromatic',
    'chromatic_first',
    'measured_l_chromatic',
    'measured_wall_l_chromatic',
    'measured_scene_l_chromatic',
    'wall_minus_scene_l_chromatic',
    'series',
]

_LUMINANCES = [
    'measured_l',
    'measured_wall_l',
    'measured_scene_l',
    'wall_minus_scene_l',
]

_AGGREGATORS = dict(
    mean=np.mean,
    std=np.std,
    min=np.min,
    max=np.max,
    range=lambda x: np.max(x) - np.min(x)
)

_LASTS = [1, 2, 3, 4, 5]


def _aggregate_last_reversals(df,
                              column='measured_l_achromatic',
                              aggregators=None,
                              num_reversals=3):
    def name(agg):
        # N.B. if studying the effect of taking last x reversals, we will need to tidy column names into a new column
        return '%s__%s__last%d' % (column, agg, num_reversals)
    if aggregators is None:
        aggregators = _AGGREGATORS
    if df.reversal.sum() < num_reversals:
        return pd.Series({name(agg_name): np.nan for agg_name in aggregators})
    return pd.Series({name(agg_name): agg(df[df.reversal][column].values[-num_reversals:])
                      for agg_name, agg in aggregators.items()})


def _munge_from_simone_excel(xlsx=LATEST_XLSX):
    """Very hacky munging of Simone's original data into tidy dataframes."""
    # To keep things simple for her, use no categoricals.

    # Open the workbook; note:
    #   http://stackoverflow.com/questions/31416842/openpyxl-does-not-close-excel-workbook-in-read-only-mode
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    assert set(wb.get_sheet_names()) == {'Results Lum Lab',
                                         'Results lum scene',
                                         'Results lum wall',
                                         'Analysis col',
                                         'INTRA-P VAR col',
                                         'L_a_b',
                                         'TABLES'}

    # --- Munge color information

    def munge_colors():
        # N.B. 'L_a_b' is hidden in the xlsx that Simone sent
        # There are also many hidden rows in there, with contradictory info...
        # So careful
        lab_sheet = wb['L_a_b']
        results_sheet = wb['Results Lum Lab']
        lum_wall_sheet = wb['Results lum wall']
        lum_scene_sheet = wb['Results lum scene']
        lab_colors = []
        for row in lab_sheet.rows:
            if row[1].value is not None and row[10].value is not None:
                name = row[10].value
                full_name = row[1].value
                if full_name.startswith(name):  # Weird stuff from Sim...
                    # A color that went into the experiment
                    # N.B. there are repeated colors, I guess we need to take into account only visible rows...
                    #      openpyxl seems to fail to read hidden properly.
                    lab_colors.append(dict(
                        name=name,
                        full_name=full_name,
                        photoshop_l=row[2].value,
                        photoshop_a=row[3].value,
                        photoshop_b=row[4].value,
                        photoshop_friday_l=row[6].value,
                        photoshop_friday_a=row[7].value,
                        photoshop_friday_b=row[8].value,
                        measured_l=row[11].value,
                        measured_a=row[12].value,
                        measured_b=row[13].value,
                    ))

        lab_df = pd.DataFrame(lab_colors, columns=['name', 'full_name',
                                                   'measured_l', 'measured_a', 'measured_b',
                                                   'photoshop_l', 'photoshop_a', 'photoshop_b'])
        # Add some info from the result xxx sheets
        colors = []
        for row in [3, 4, 5] + range(8, 37):
            name = results_sheet.cell(row=row, column=1).value
            measured_luminance = results_sheet.cell(row=row, column=2).value
            measured_wall_luminance = lum_wall_sheet.cell(row=row, column=2).value
            measured_scene_luminance = lum_scene_sheet.cell(row=row, column=2).value
            colors.append(dict(
                name=name,
                luminance=measured_luminance,  # will use this to sanity check
                measured_wall_l=measured_wall_luminance,
                measured_scene_l=measured_scene_luminance,
                achromatic=name.startswith('M'),
                step=None
            ))
        colors_df = pd.DataFrame(colors)
        steps = colors_df.query('achromatic').luminance.diff()
        colors_df.loc[colors_df.achromatic, 'step'] = steps
        colors_df.loc[colors_df.achromatic, 'step_index'] = np.arange(len(steps))
        colors_df = pd.merge(colors_df, lab_df, on='name')

        # Sanity checks
        assert colors_df.name.nunique() == len(colors_df)
        assert (len(colors_df) == 32)
        assert ((~colors_df.achromatic).sum() == 3)
        assert (colors_df.achromatic.sum() == 29)
        staircase = colors_df.query('achromatic')
        lum_range = staircase.luminance.max() - staircase.luminance.min()
        assert abs(lum_range - 29.644179823) - 1E-9
        assert np.allclose(colors_df.luminance, colors_df.measured_l)

        colors_df = colors_df[['name', 'full_name', 'achromatic', 'step_index', 'step',
                               'measured_l', 'measured_a', 'measured_b',
                               'photoshop_l', 'photoshop_a', 'photoshop_b',
                               'measured_wall_l', 'measured_scene_l']].set_index('name')
        colors_df['wall_minus_scene_l'] = colors_df.measured_wall_l - colors_df.measured_scene_l
        return colors_df

    colors_df = munge_colors()

    # --- Munge participants information

    def munge_participants(num_participants=35):
        subjects_sheet = wb['Analysis col']
        #
        # Adds subject X (also known as Mr. C)
        # TODO: add actual data and data for simple reversals. From Sim:
        # ----------------------
        # - The other thing is that for participant 0 I have participant and experiment
        #   time details as well and data for staircase with single reversals
        #   (personal details in cells colum CS of sheet 'Analysis Col' and Data for
        #    staircase with single reversals in columns CP & CQ for the same sheet ).
        #   It's probably good to include, not for the main analysis but for a comparsion
        #   of a result from an trained observer.
        # ----------------------
        #
        subjects = [dict(
            subject_id=0,
            omit=True,
            exp_start=None,
            age=None,
            gender=None,
            education=None,
            vision=None,
        )]

        def must_omit(omit):
            return omit is not None and omit.startswith('OMIT')

        for column in range(11, num_participants * 2 + 10, 2):
            date = subjects_sheet.cell(row=3, column=column).value
            time = subjects_sheet.cell(row=5, column=column).value
            if isinstance(time, string_types) and ';' in time:
                hour, minute = map(int, time.split(';'))
                time = datetime.time(hour=hour, minute=minute)
            subjects.append(dict(
                subject_id=subjects_sheet.cell(row=1, column=column).value,
                omit=must_omit(subjects_sheet.cell(row=2, column=column).value),
                exp_start=datetime.datetime.combine(date, time),
                age=subjects_sheet.cell(row=7, column=column).value,
                gender=subjects_sheet.cell(row=9, column=column).value,
                education=subjects_sheet.cell(row=11, column=column).value,
                vision=subjects_sheet.cell(row=14, column=column).value,
            ))

        return pd.DataFrame(subjects)[['subject_id',
                                       'omit',
                                       'exp_start',
                                       'age',
                                       'gender',
                                       'education',
                                       'vision']].set_index('subject_id')
    participants_df = munge_participants()

    # --- Munge trials information

    def munge_trials():
        results_sheet = wb['Results Lum Lab']
        response_matcher = re.compile(r'(.*)-(.*):(.*)')
        response_columns = [(7, 0), (8, 0), (9, 0), (10, 0), (11, 0), (12, 0),
                            (15, 1), (16, 1), (17, 1), (18, 1), (19, 1), (20, 1)]
        current_participant = 0
        current_step_index = 0
        current_reversal_method = 'simple-reversals'
        responses = []

        def parse_response(response, staircase_repetition, step_index):
            try:
                color1, color2, response = map(lambda x: x.strip(),
                                               response_matcher.match(response).groups())
                staircase = '%s-%s' % (
                    'M' if color1.startswith('M') else color1,
                    'M' if color2.startswith('M') else color2
                )
                more_light = color1 if int(response[0]) == 1 else color2

                response = dict(
                    participant=current_participant,
                    staircase=staircase,
                    staircase_repetition=staircase_repetition,
                    step_index=step_index,
                    color1=color1,
                    color2=color2,
                    response=more_light,
                    reversal_method=current_reversal_method,
                )

                return response
            except (TypeError, AttributeError):
                return None

        for row in results_sheet.rows:
            # New participant?
            participant = row[6].value
            if isinstance(participant, string_types):
                if participant.lower().startswith('participant'):
                    current_participant = int(participant.split()[-1])
                    current_step_index = 0
                elif participant.lower() == 'method 2':
                    current_reversal_method = 'double-reversals'
            # Add new observations?
            responses_row = [parse_response(row[column].value,
                                            staircase_repetition=repetition,
                                            step_index=current_step_index)
                             for column, repetition in response_columns]
            if any(responses_row):
                responses += [obs for obs in responses_row if obs is not None]
                current_step_index += 1

        trials_df = pd.DataFrame(responses)
        # Unnormalise a bit more
        trials_df['achromatic'] = trials_df.apply(lambda row: row.color1 if row.color1.startswith('M') else row.color2,
                                                  axis=1)
        trials_df['chromatic'] = trials_df.apply(lambda row: row.color2 if row.color1.startswith('M') else row.color1,
                                                 axis=1)
        trials_df['chromatic_first'] = trials_df.color1.str.startswith('C')
        trials_df['chromatic_brighter'] = trials_df.response.str.startswith('C')

        # Add reversal and reversal_count columns
        trials = []
        for _, obs in trials_df.groupby(['participant', 'staircase', 'staircase_repetition']):
            obs = obs.copy().reset_index(drop=True)

            def repetition(row, repetition_counter=Counter()):
                key = (row.color1, row.color2)
                repetition_counter[key] += 1
                return repetition_counter[key]

            obs = obs.sort_values('step_index')
            obs['reversal'] = obs.chromatic_brighter[1:] != obs.chromatic_brighter[:-1]
            obs.set_value(0, 'reversal', False)
            obs['pair_count'] = obs.apply(repetition, axis=1)
            trials.append(obs)
        trials_df = pd.concat(trials, ignore_index=True).reset_index(drop=True)
        trials_df = trials_df.sort_values(['participant',
                                           'staircase',
                                           'staircase_repetition',
                                           'reversal_method',
                                           'step_index',
                                           ]).reset_index(drop=True)
        return trials_df[['participant', 'reversal_method',
                          'staircase', 'staircase_repetition',
                          'step_index',
                          'color1', 'color2', 'pair_count',
                          'chromatic_first', 'achromatic', 'chromatic',
                          'response', 'chromatic_brighter',
                          'reversal']]

    trials_df = munge_trials()

    # Having luminance in the trials dataframe will prove useful
    # http://stackoverflow.com/questions/20206615/how-can-a-pandas-merge-preserve-order
    num_trials = len(trials_df)
    trials_df = (trials_df.
                 merge(colors_df[['measured_l', 'measured_a', 'measured_b',
                                  'measured_wall_l', 'measured_scene_l', 'wall_minus_scene_l']],
                       left_on='achromatic', right_index=True,
                       how='left').
                 sort_index())
    trials_df = (trials_df.
                 merge(colors_df[['measured_l', 'measured_a', 'measured_b',
                                  'measured_wall_l', 'measured_scene_l', 'wall_minus_scene_l']],
                       left_on='chromatic', right_index=True,
                       suffixes=['_achromatic', '_chromatic'],
                       how='left').
                 sort_index())
    assert len(trials_df) == num_trials

    # Let's create an "experiments" dataframe with each experiment in a row.
    # The terms "experiment", "trial"... are a bit clunky.
    # Simone's has decided to call "trial" to each "brighter" decision, so lets keep it like that
    exp_key = ['participant',
               'staircase',
               'staircase_repetition',
               'reversal_method']
    exps_df = _trials_to_exp(trials_df, exp_attributes=exp_key)
    exps_df = exps_df.merge(participants_df, how='left',
                            left_on='participant', right_index=True)
    exps_df = exps_df.sort_values(exp_key)

    # Denormalise also chromatic info in the exps dataframe
    def infer_chromatic(staircase_id):
        color1, color2 = staircase_id.split('-')
        return color1 if color1.startswith('C') else color2
    exps_df['chromatic'] = exps_df.staircase.apply(infer_chromatic)
    exps_df['chromatic_first'] = exps_df.staircase.str.startswith('C')
    exps_df = exps_df[[col for col in exps_df.columns if col != 'series'] + ['series']]

    # Add here summaries for the staircases, so Sim does not need to worry about these computations
    exps_df = exps_df.merge(colors_df[_LUMINANCES], left_on='chromatic', right_index=True)
    exps_df = exps_df.rename(columns={col: col + '_chromatic' for col in _LUMINANCES})
    aggs = [partial(_aggregate_last_reversals, num_reversals=num_reversals, column=column + '_achromatic')
            for num_reversals, column in product(_LASTS, _LUMINANCES)]
    for agg in aggs:
        exps_df = exps_df.merge(exps_df.series.apply(agg), left_index=True, right_index=True)
    exps_df['staircase_length'] = exps_df.series.apply(len)

    # Reorder slightly the columns, so "series" come right before the staircase summaries
    staircases_summaries = [col for col in exps_df.columns if col not in EXPS_NON_STAIRCASE_COLUMNS]
    exps_df = exps_df[EXPS_NON_STAIRCASE_COLUMNS + staircases_summaries]

    # N.B. exp_start is wrong for method 2 (double revelsals);
    # Sim should put "start" on top of experiments, it is not a user attr.

    # Cache, html
    trials_df.to_html(op.join(DATA_DIR, 'trials.html'))
    colors_df.to_html(op.join(DATA_DIR, 'colors.html'))
    participants_df.to_html(op.join(DATA_DIR, 'participants.html'))
    exps_df.drop('series', axis=1).to_html(op.join(DATA_DIR, 'experiments.html'))
    # Cache, excel
    trials_df.to_excel(op.join(DATA_DIR, 'trials.xlsx'))
    colors_df.to_excel(op.join(DATA_DIR, 'colors.xlsx'))
    participants_df.to_excel(op.join(DATA_DIR, 'participants.xlsx'))
    exps_df.drop('series', axis=1).to_excel(op.join(DATA_DIR, 'experiments.xlsx'))
    # Cache, pickle (here we do save the useful "series" column in the experiments table)
    trials_df.to_pickle(op.join(DATA_DIR, 'trials.pkl'))
    colors_df.to_pickle(op.join(DATA_DIR, 'colors.pkl'))
    participants_df.to_pickle(op.join(DATA_DIR, 'participants.pkl'))
    exps_df.to_pickle(op.join(DATA_DIR, 'experiments.pkl'))

    return colors_df, participants_df, trials_df, exps_df


def staircases(recache=False):
    """
    Returns 4 dataframes with all the data from the experiments:
      - colors: each row contains a color information
      - participants: each row contains a participant information
      - trials: each row contains a trial (here a single decision of which color is brighter)
      - experiments: each row is an experiment;
        contains a special column, "series", with all the trials of the experiment sorted

    Parameters
    ----------
    recache : bool, default False
      If True, the data is read from the original excel files and cached to files.
      If False, the data caches are tried first.
    """
    try:
        if recache:
            raise Exception()
        colors_df = pd.read_pickle(op.join(DATA_DIR, 'colors.pkl'))
        participants_df = pd.read_pickle(op.join(DATA_DIR, 'participants.pkl'))
        trials_df = pd.read_pickle(op.join(DATA_DIR, 'trials.pkl'))
        exps_df = pd.read_pickle(op.join(DATA_DIR, 'experiments.pkl'))
        return colors_df, participants_df, trials_df, exps_df
    except Exception:
        return _munge_from_simone_excel()


if __name__ == '__main__':
    import argh
    argh.dispatch_command(staircases)
